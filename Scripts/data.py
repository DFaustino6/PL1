import os.path
from os import path
import openpyxl as op
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import re
sheet = 'Threaded'
workbook = 'gui_data.xlsx'
directory = 'results/gui/ThreadedTests/'

def write_excel(results):    
    if not path.exists(workbook):
        wb = Workbook(workbook)
        
    else:
        wb = op.load_workbook(workbook)
    
    if sheet not in wb.sheetnames:
        ws = wb.create_sheet(sheet)
    else:
        ws = wb[sheet]
    
    wb.save(workbook)

    for i in range(34):#atribuir o numero de teste a cada coluna
            ws.cell(row = i+2,column = 1, value = '#'+str(i+1))

    ws.cell(row=i+1,column = 1, value = 'Média')
    ws.cell(row=i+2,column = 1, value = 'DPadrão')

    #j = 2;
    #for test in results:#identificar a que quadrado pertence os testes
    #    value = test
    #    ws.cell(row = 1,column = j, value = value)
    #    j+=1;
    
    #distribui os dados dos ficheiros, um ficheiro por cada linha
    row = 2;
    j = 2
    type = ''
    for file in results:
        f = open(directory+file,'r')
        values = f.readlines()
        col=2   
        total = 0
        ws.cell(row = 1,column = j, value = file.partition('_')[0][1:]+"x"+file.partition('_')[0][1:]+"_th")
        j+=1
        for v in values[1:]:     
            if v != '\n':  
                #print(re.findall(r"\d*\,\d+|\d+",v)[0])  
                number = re.findall(r"\d*\,\d+|\d+",v)[0]
                ws.cell(row = col,column = row, value = float(number.replace(',','.')))
                col+=1
                
                total += float(number.replace(',','.'))
        
        ws.cell(row=col,column=row,value=(total/(len(values)-1))*pow(1,-3))

        type = file[0:1]
        row+=1    

    wb.save(workbook)
    wb.close()



def main():
    results = os.listdir(directory)
    sort_by_type = sorted(results, key=lambda x:  (x[0:1],int(x[1:].partition('_')[0])))
    #for item in sort_by_type:
    #    print(item)
    write_excel(sort_by_type)

if __name__ == "__main__":
	main()
